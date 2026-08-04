"""Ingest machine documentation into the Neo4j document domain.

This is the active replacement for the older alternate/ prototype. It writes
documentation into Neo4j, not SINDIT/GraphDB, and keeps documents partitioned
by usecase via explicit `usecase` properties on every node.

Graph shape:

  (:DocumentSource {id, usecase, source})
      -[:HAS_FILE]->
  (:DocumentFile {id, relative_path, usecase, machine, ...})
      -[:HAS_CHUNK {index}]->
  (:Document {id, text, page, chunk_index, embedding, usecase, ...})
      -[:NEXT_CHUNK]->(:Document)
      -[:MENTIONS]->(:Entity {id, name, type, usecase, embedding, ...})

  (:Entity)-[:REL {type, source_chunk_id}]->(:Entity)

This gives the retriever both chunk-level vector retrieval and a grounded
semantic layer whose entities and relations always trace back to document
chunks.
"""

from __future__ import annotations

import asyncio
import concurrent.futures
import hashlib
import logging
import re
import zipfile
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Iterator, Optional, Sequence
from xml.etree import ElementTree as ET

from backend.agents.config import (
    DATA_DIR,
    NEO4J_CONNECT_TIMEOUT_S,
    NEO4J_DATABASE,
    NEO4J_PASSWORD,
    NEO4J_URI,
    NEO4J_USERNAME,
    SINDIT_API_URL,
    SINDIT_ENABLED,
    SINDIT_TIMEOUT_S,
)
from backend.agents.llm.entity_canonicalizer import EntityCanonicalizer
from backend.agents.llm.entity_extractor import EntityExtractor, normalize_entity_name
from backend.agents.usecase import USECASE_GENERIC, resolve_usecase

logger = logging.getLogger(__name__)

_DOC_SCHEMA_VERSION = 3
_DEFAULT_CHUNK_CHARS = 3500
_DEFAULT_CHUNK_OVERLAP = 400
_SUPPORTED_SUFFIXES = {".pdf", ".md", ".txt", ".xlsx", ".docx", ".ods"}
_SEMANTIC_SUFFIXES = {".pdf", ".md", ".docx"}
_ENTITY_VECTOR_INDEX = "entity_vector_index"
_RESOLVABLE_ASSET_ENTITY_TYPES = frozenset({"Machine", "Tool", "Component", "Subsystem"})
_BLANK_PAGE_HINTS = (
    "this page intentionally left blank",
    "página intencionadamente en blanco",
    "página en blanco",
    "blank page",
)


@dataclass(frozen=True)
class TextUnit:
    text: str
    page: Optional[int] = None
    section: Optional[str] = None


@dataclass(frozen=True)
class DocumentChunk:
    id: str
    file_id: str
    source_id: str
    usecase: str
    source: str
    dataset_id: str
    file_name: str
    relative_path: str
    subdirectory: Optional[str]
    text: str
    page: Optional[int]
    section: Optional[str]
    chunk_index: int
    document_type: str
    machine: Optional[str]
    machine_ids: list[str]
    machine_uri: Optional[str]
    language_code: Optional[str]
    original_language: Optional[str]
    updated_at: str
    embedding: Optional[list[float]] = None


@dataclass(frozen=True)
class DocumentFileRecord:
    id: str
    source_id: str
    usecase: str
    source: str
    dataset_id: str
    file_name: str
    relative_path: str
    subdirectory: Optional[str]
    document_type: str
    machine: Optional[str]
    machine_ids: list[str]
    machine_uri: Optional[str]
    language_code: Optional[str]
    original_language: Optional[str]
    page_count: int
    chunk_count: int
    updated_at: str


@dataclass(frozen=True)
class SemanticEntityRecord:
    id: str
    usecase: str
    type: str
    name: str
    name_norm: str
    aliases: list[str]
    description: str
    updated_at: str
    canonical_id: Optional[str] = None
    embedding: Optional[list[float]] = None


@dataclass(frozen=True)
class SemanticMentionRecord:
    chunk_id: str
    entity_id: str
    confidence: float
    updated_at: str


@dataclass(frozen=True)
class SemanticRelationRecord:
    source_entity_id: str
    target_entity_id: str
    relation_type: str
    source_chunk_id: str
    confidence: float
    updated_at: str


@dataclass(frozen=True)
class SemanticGraphRecords:
    entities: list[SemanticEntityRecord]
    mentions: list[SemanticMentionRecord]
    relations: list[SemanticRelationRecord]
    extractor_failures: int = 0
    new_entity_count: int = 0


@dataclass
class IngestSummary:
    files_seen: int = 0
    files_parsed: int = 0
    files_ingested: int = 0
    chunks_created: int = 0
    entities_extracted: int = 0
    relations_extracted: int = 0
    extractor_failures: int = 0
    skipped_unsupported: int = 0
    skipped_empty: int = 0
    warnings: list[str] = field(default_factory=list)


class _DocumentEntityCanonicalIdResolver:
    def __init__(self, assets: Sequence[dict[str, Any]]) -> None:
        self._assets_by_key: dict[str, set[str]] = {}
        for asset in assets:
            iri = str(asset.get("iri") or "").strip()
            if not iri:
                continue
            label = str(asset.get("label") or "").strip()
            for surface in {label, _asset_tail(iri)}:
                for key in _entity_surface_keys(surface):
                    self._assets_by_key.setdefault(key, set()).add(iri)

    def resolve(
        self,
        *,
        name: str,
        entity_type: str,
        aliases: Sequence[str],
        usecase: str,
        machine_hint: Optional[str] = None,
        machine_uri: Optional[str] = None,
    ) -> Optional[str]:
        del usecase
        surface_keys = {
            key
            for surface in [str(name or "").strip(), *(str(alias or "").strip() for alias in aliases)]
            if str(surface or "").strip()
            for key in _entity_surface_keys(surface)
        }

        if entity_type == "Machine" and machine_uri:
            machine_keys = _entity_surface_keys(machine_hint) | _entity_surface_keys(_asset_tail(machine_uri))
            if surface_keys & machine_keys:
                return str(machine_uri).strip() or None

        if entity_type not in _RESOLVABLE_ASSET_ENTITY_TYPES:
            return None

        candidates = {
            iri
            for surface_key in surface_keys
            for iri in self._assets_by_key.get(surface_key, set())
        }
        if len(candidates) == 1:
            return next(iter(candidates))
        return None


def ingest_machinedocs(
    *,
    root: Path | str | None = None,
    usecases: Optional[Sequence[str]] = None,
    dry_run: bool = False,
    clear_documents: bool = False,
    embed_documents: bool = True,
    extract_entities: bool = False,
    entity_extractor_model: Optional[str] = None,
    limit_files: Optional[int] = None,
    chunk_chars: int = _DEFAULT_CHUNK_CHARS,
    chunk_overlap: int = _DEFAULT_CHUNK_OVERLAP,
    uri: str = NEO4J_URI,
    username: str = NEO4J_USERNAME,
    password: str = NEO4J_PASSWORD,
    database: str = NEO4J_DATABASE,
    connect_timeout_s: float = NEO4J_CONNECT_TIMEOUT_S,
) -> IngestSummary:
    docs_root = Path(root) if root is not None else Path(DATA_DIR) / "machinedocs"
    summary = IngestSummary()
    selected_usecases = {
        normalized
        for value in (usecases or [])
        if (normalized := resolve_usecase(usecase=value, fallback_generic=False))
    }

    parsed_files: list[tuple[DocumentFileRecord, list[DocumentChunk], Optional[SemanticGraphRecords]]] = []
    extractor = EntityExtractor(enabled=extract_entities, model=entity_extractor_model)
    semantic_canonicalizers: dict[str, EntityCanonicalizer] = {}
    canonical_id_resolver = _build_document_entity_canonical_id_resolver(summary.warnings) if extract_entities else None
    if extract_entities and not extractor.is_enabled():
        summary.warnings.append(
            "Entity extraction requested but the Groq-backed extractor is unavailable; semantic preview skipped."
        )
    for path in iter_document_paths(docs_root):
        summary.files_seen += 1
        file_usecase = _usecase_for_path(path, docs_root)
        if selected_usecases and file_usecase not in selected_usecases:
            continue
        try:
            parsed = parse_document_file(
                path,
                root=docs_root,
                chunk_chars=chunk_chars,
                chunk_overlap=chunk_overlap,
                embed=embed_documents and not dry_run,
            )
        except Exception as exc:
            relative = path.relative_to(docs_root).as_posix()
            warning = f"Failed to parse {relative}: {exc}"
            summary.warnings.append(warning)
            logger.warning(warning)
            continue
        if parsed is None:
            summary.skipped_empty += 1
            continue
        file_record, chunks = parsed
        summary.files_parsed += 1
        semantic_records: Optional[SemanticGraphRecords] = None
        if extract_entities and path.suffix.lower() in _SEMANTIC_SUFFIXES and extractor.is_enabled():
            canonicalizer = semantic_canonicalizers.setdefault(
                file_record.usecase,
                EntityCanonicalizer(
                    usecase=file_record.usecase,
                    canonical_id_resolver=canonical_id_resolver.resolve if canonical_id_resolver is not None else None,
                ),
            )
            semantic_records = _build_semantic_graph_records(
                chunks,
                extractor=extractor,
                canonicalizer=canonicalizer,
                attach_embeddings=not dry_run,
            )
            summary.entities_extracted += semantic_records.new_entity_count
            summary.relations_extracted += len(semantic_records.relations)
            summary.extractor_failures += semantic_records.extractor_failures
        parsed_files.append((file_record, chunks, semantic_records))
        if limit_files is not None and len(parsed_files) >= int(limit_files):
            break

    if dry_run:
        summary.files_ingested = 0
        summary.chunks_created = sum(len(chunks) for _, chunks, _ in parsed_files)
        return summary

    store = DocumentGraphStore(
        uri=uri,
        username=username,
        password=password,
        database=database,
        connect_timeout_s=connect_timeout_s,
    )
    try:
        if clear_documents:
            target_usecase = next(iter(selected_usecases)) if len(selected_usecases) == 1 else None
            store.clear_document_domain(usecase=target_usecase)

        for file_record, chunks, semantic_records in parsed_files:
            store.upsert_file(file_record, chunks, semantic_records=semantic_records)
            summary.files_ingested += 1
            summary.chunks_created += len(chunks)
    finally:
        store.close()

    return summary


def _preview_semantic_extraction(
    chunks: Sequence[DocumentChunk],
    *,
    extractor: EntityExtractor,
    canonicalizer: EntityCanonicalizer,
) -> tuple[int, int, int]:
    records = _build_semantic_graph_records(
        chunks,
        extractor=extractor,
        canonicalizer=canonicalizer,
        attach_embeddings=False,
    )
    return records.new_entity_count, len(records.relations), records.extractor_failures


def _build_semantic_graph_records(
    chunks: Sequence[DocumentChunk],
    *,
    extractor: EntityExtractor,
    canonicalizer: EntityCanonicalizer,
    attach_embeddings: bool,
) -> SemanticGraphRecords:
    initial_count = len(canonicalizer.list_entities())
    entity_rows: dict[str, SemanticEntityRecord] = {}
    mention_rows: dict[tuple[str, str], SemanticMentionRecord] = {}
    relation_rows: dict[tuple[str, str, str, str], SemanticRelationRecord] = {}
    extractor_failures = 0

    for chunk in chunks:
        result = extractor.extract_from_chunk(
            chunk.text,
            usecase=chunk.usecase,
            machine_hint=chunk.machine,
            source_hint=chunk.file_name,
        )
        if any(warning.startswith("extractor_error:") for warning in result.warnings):
            extractor_failures += 1

        chunk_entities: dict[tuple[str, str], str] = {}
        for entity in result.entities:
            canonical = canonicalizer.register(
                name=entity.name,
                entity_type=entity.type,
                aliases=entity.aliases,
                machine_hint=chunk.machine,
                machine_uri=chunk.machine_uri,
            )
            entity_rows[canonical.id] = SemanticEntityRecord(
                id=canonical.id,
                usecase=canonical.usecase,
                type=canonical.type,
                name=canonical.name,
                name_norm=canonical.name_norm,
                aliases=list(canonical.aliases),
                description=entity_rows.get(canonical.id, SemanticEntityRecord(
                    id=canonical.id,
                    usecase=canonical.usecase,
                    type=canonical.type,
                    name=canonical.name,
                    name_norm=canonical.name_norm,
                    aliases=list(canonical.aliases),
                    description=_entity_description_from_chunk(chunk.text),
                    updated_at=chunk.updated_at,
                    canonical_id=canonical.canonical_id,
                )).description,
                updated_at=chunk.updated_at,
                canonical_id=canonical.canonical_id,
                embedding=entity_rows.get(canonical.id).embedding if canonical.id in entity_rows else None,
            )
            chunk_entities[(canonical.type, canonical.name_norm)] = canonical.id
            mention_rows[(chunk.id, canonical.id)] = SemanticMentionRecord(
                chunk_id=chunk.id,
                entity_id=canonical.id,
                confidence=1.0,
                updated_at=chunk.updated_at,
            )

        for relation in result.relations:
            source_key = (relation.src_type, normalize_entity_name(relation.src_name))
            target_key = (relation.dst_type, normalize_entity_name(relation.dst_name))
            source_entity_id = chunk_entities.get(source_key)
            target_entity_id = chunk_entities.get(target_key)
            if not source_entity_id or not target_entity_id:
                continue
            relation_rows[(source_entity_id, target_entity_id, relation.rel_type, chunk.id)] = SemanticRelationRecord(
                source_entity_id=source_entity_id,
                target_entity_id=target_entity_id,
                relation_type=relation.rel_type,
                source_chunk_id=chunk.id,
                confidence=relation.confidence,
                updated_at=chunk.updated_at,
            )

    entities = sorted(entity_rows.values(), key=lambda item: (item.type, item.name_norm, item.id))
    if attach_embeddings:
        _attach_entity_embeddings(entities)
    return SemanticGraphRecords(
        entities=entities,
        mentions=list(mention_rows.values()),
        relations=list(relation_rows.values()),
        extractor_failures=extractor_failures,
        new_entity_count=max(0, len(canonicalizer.list_entities()) - initial_count),
    )


def iter_document_paths(root: Path) -> Iterator[Path]:
    if not root.exists():
        raise FileNotFoundError(root)
    for path in sorted(root.rglob("*")):
        if not path.is_file():
            continue
        if ":Zone.Identifier" in path.name:
            continue
        if path.suffix.lower() not in _SUPPORTED_SUFFIXES:
            continue
        yield path


def parse_document_file(
    path: Path,
    *,
    root: Path,
    chunk_chars: int = _DEFAULT_CHUNK_CHARS,
    chunk_overlap: int = _DEFAULT_CHUNK_OVERLAP,
    embed: bool = True,
) -> Optional[tuple[DocumentFileRecord, list[DocumentChunk]]]:
    usecase = _usecase_for_path(path, root)
    relative_path = path.relative_to(root).as_posix()
    source = usecase.lower()
    dataset_id = source
    subdirectory = _subdirectory_for_path(path, root)
    document_type = _infer_document_type(path)
    machine_ids = _infer_machine_ids(path, usecase)
    machine = ",".join(machine_ids) if machine_ids else None
    machine_uri = _machine_uri(machine_ids)

    text_units = _extract_text_units(path)
    page_count = sum(1 for unit in text_units if unit.page is not None)
    combined_sample = "\n".join(unit.text[:600] for unit in text_units[:3])
    language_code, original_language = _guess_language(path, combined_sample)

    source_id = _stable_id("docsource", usecase)
    file_id = _stable_id("docfile", relative_path)
    updated_at = _now_iso()

    chunks: list[DocumentChunk] = []
    chunk_index = 0
    for unit in text_units:
        for chunk_text in _chunk_text(unit.text, chunk_chars=chunk_chars, chunk_overlap=chunk_overlap):
            if len(chunk_text.strip()) < 40:
                continue
            chunks.append(
                DocumentChunk(
                    id=_stable_id("doc", f"{relative_path}:{unit.page}:{chunk_index}"),
                    file_id=file_id,
                    source_id=source_id,
                    usecase=usecase,
                    source=source,
                    dataset_id=dataset_id,
                    file_name=path.name,
                    relative_path=relative_path,
                    subdirectory=subdirectory,
                    text=chunk_text,
                    page=unit.page,
                    section=unit.section,
                    chunk_index=chunk_index,
                    document_type=document_type,
                    machine=machine,
                    machine_ids=list(machine_ids),
                    machine_uri=machine_uri,
                    language_code=language_code,
                    original_language=original_language,
                    updated_at=updated_at,
                )
            )
            chunk_index += 1

    if not chunks:
        return None

    if embed:
        _attach_embeddings(chunks)

    file_record = DocumentFileRecord(
        id=file_id,
        source_id=source_id,
        usecase=usecase,
        source=source,
        dataset_id=dataset_id,
        file_name=path.name,
        relative_path=relative_path,
        subdirectory=subdirectory,
        document_type=document_type,
        machine=machine,
        machine_ids=list(machine_ids),
        machine_uri=machine_uri,
        language_code=language_code,
        original_language=original_language,
        page_count=page_count,
        chunk_count=len(chunks),
        updated_at=updated_at,
    )
    return file_record, chunks


class DocumentGraphStore:
    def __init__(
        self,
        *,
        uri: str = NEO4J_URI,
        username: str = NEO4J_USERNAME,
        password: str = NEO4J_PASSWORD,
        database: str = NEO4J_DATABASE,
        connect_timeout_s: float = NEO4J_CONNECT_TIMEOUT_S,
    ) -> None:
        from neo4j import GraphDatabase  # type: ignore[import-untyped]

        self._driver = GraphDatabase.driver(
            uri,
            auth=(username, password),
            connection_timeout=connect_timeout_s,
            connection_acquisition_timeout=connect_timeout_s,
        )
        self._database = database
        self._ensure_schema()

    def close(self) -> None:
        self._driver.close()

    def clear_document_domain(self, *, usecase: Optional[str] = None) -> None:
        with self._driver.session(database=self._database) as session:
            session.run(
                """
                MATCH (n)
                WHERE (n:Document OR n:DocumentFile OR n:DocumentSource OR n:Entity)
                  AND ($usecase IS NULL OR n.usecase = $usecase)
                DETACH DELETE n
                """,
                usecase=usecase,
            ).consume()

    def upsert_file(
        self,
        file_record: DocumentFileRecord,
        chunks: Sequence[DocumentChunk],
        semantic_records: Optional[SemanticGraphRecords] = None,
    ) -> None:
        with self._driver.session(database=self._database) as session:
            session.execute_write(self._tx_upsert_file, file_record, list(chunks), semantic_records)

    def _ensure_schema(self) -> None:
        with self._driver.session(database=self._database) as session:
            session.run(
                "CREATE CONSTRAINT doc_source_id_unique IF NOT EXISTS "
                "FOR (s:DocumentSource) REQUIRE s.id IS UNIQUE"
            )
            session.run(
                "CREATE CONSTRAINT doc_file_id_unique IF NOT EXISTS "
                "FOR (f:DocumentFile) REQUIRE f.id IS UNIQUE"
            )
            session.run(
                "CREATE CONSTRAINT document_id_unique IF NOT EXISTS "
                "FOR (d:Document) REQUIRE d.id IS UNIQUE"
            )
            session.run(
                "CREATE INDEX doc_source_idx IF NOT EXISTS "
                "FOR (d:Document) ON (d.source)"
            )
            session.run(
                "CREATE INDEX doc_usecase_idx IF NOT EXISTS "
                "FOR (d:Document) ON (d.usecase)"
            )
            session.run(
                "CREATE INDEX doc_file_usecase_idx IF NOT EXISTS "
                "FOR (f:DocumentFile) ON (f.usecase)"
            )
            session.run(
                "CREATE CONSTRAINT entity_id_unique IF NOT EXISTS "
                "FOR (e:Entity) REQUIRE e.id IS UNIQUE"
            )
            session.run(
                "CREATE INDEX entity_usecase_idx IF NOT EXISTS "
                "FOR (e:Entity) ON (e.usecase)"
            )
            session.run(
                "CREATE INDEX entity_type_idx IF NOT EXISTS "
                "FOR (e:Entity) ON (e.type)"
            )
            session.run(
                "CREATE INDEX entity_name_idx IF NOT EXISTS "
                "FOR (e:Entity) ON (e.name_norm)"
            )
            session.run(
                "CREATE INDEX entity_canonical_id_idx IF NOT EXISTS "
                "FOR (e:Entity) ON (e.canonical_id)"
            )
            try:
                session.run(
                    "CREATE VECTOR INDEX doc_vector_index IF NOT EXISTS "
                    "FOR (d:Document) ON (d.embedding) "
                    "OPTIONS {indexConfig: {"
                    "  `vector.dimensions`: 384,"
                    "  `vector.similarity_function`: 'cosine'"
                    "}}"
                )
            except Exception:
                logger.debug("Document vector index creation skipped")
            try:
                session.run(
                    f"CREATE VECTOR INDEX {_ENTITY_VECTOR_INDEX} IF NOT EXISTS "
                    "FOR (e:Entity) ON (e.embedding) "
                    "OPTIONS {indexConfig: {"
                    "  `vector.dimensions`: 384,"
                    "  `vector.similarity_function`: 'cosine'"
                    "}}"
                )
            except Exception:
                logger.debug("Entity vector index creation skipped")
            session.run(
                "MERGE (sv:SchemaVersion {domain: 'documents'}) "
                "SET sv.version = $version, sv.updated_at = $ts",
                version=_DOC_SCHEMA_VERSION,
                ts=_now_iso(),
            )

    @staticmethod
    def _tx_upsert_file(
        tx: Any,
        file_record: DocumentFileRecord,
        chunks: list[DocumentChunk],
        semantic_records: Optional[SemanticGraphRecords],
    ) -> None:
        tx.run(
            "MERGE (s:DocumentSource {id: $id}) SET s += $props",
            id=file_record.source_id,
            props={
                "id": file_record.source_id,
                "usecase": file_record.usecase,
                "source": file_record.source,
                "dataset_id": file_record.dataset_id,
                "label": file_record.usecase,
                "updated_at": file_record.updated_at,
            },
        )
        tx.run(
            "MERGE (f:DocumentFile {id: $id}) SET f += $props",
            id=file_record.id,
            props={
                "id": file_record.id,
                "usecase": file_record.usecase,
                "source": file_record.source,
                "dataset_id": file_record.dataset_id,
                "file_name": file_record.file_name,
                "relative_path": file_record.relative_path,
                "subdirectory": file_record.subdirectory,
                "document_type": file_record.document_type,
                "machine": file_record.machine,
                "machine_ids": file_record.machine_ids,
                "machine_uri": file_record.machine_uri,
                "language_code": file_record.language_code,
                "original_language": file_record.original_language,
                "page_count": file_record.page_count,
                "chunk_count": file_record.chunk_count,
                "updated_at": file_record.updated_at,
            },
        )
        tx.run(
            "MATCH (s:DocumentSource {id: $source_id}) "
            "MATCH (f:DocumentFile {id: $file_id}) "
            "MERGE (s)-[:HAS_FILE]->(f)",
            source_id=file_record.source_id,
            file_id=file_record.id,
        )
        tx.run(
            """
            MATCH (f:DocumentFile {id: $file_id})-[:HAS_CHUNK]->(d:Document)
            WITH collect(d.id) AS chunk_ids
            MATCH ()-[r:REL]->()
            WHERE r.source_chunk_id IN chunk_ids
            DELETE r
            """,
            file_id=file_record.id,
        )
        tx.run(
            "MATCH (f:DocumentFile {id: $file_id})-[:HAS_CHUNK]->(d:Document) "
            "DETACH DELETE d",
            file_id=file_record.id,
        )

        chunk_rows = []
        for chunk in chunks:
            chunk_rows.append(
                {
                    "id": chunk.id,
                    "file_id": chunk.file_id,
                    "chunk_index": chunk.chunk_index,
                    "props": {
                        "id": chunk.id,
                        "file_id": chunk.file_id,
                        "usecase": chunk.usecase,
                        "source": chunk.source,
                        "dataset_id": chunk.dataset_id,
                        "file_name": chunk.file_name,
                        "relative_path": chunk.relative_path,
                        "subdirectory": chunk.subdirectory,
                        "text": chunk.text,
                        "page": chunk.page,
                        "section": chunk.section,
                        "chunk_index": chunk.chunk_index,
                        "document_type": chunk.document_type,
                        "machine": chunk.machine,
                        "machine_ids": chunk.machine_ids,
                        "machine_uri": chunk.machine_uri,
                        "language_code": chunk.language_code,
                        "original_language": chunk.original_language,
                        "updated_at": chunk.updated_at,
                        "embedding": chunk.embedding,
                    },
                }
            )

        tx.run(
            """
            MATCH (f:DocumentFile {id: $file_id})
            UNWIND $rows AS row
            MERGE (d:Document {id: row.id})
            SET d += row.props
            MERGE (f)-[r:HAS_CHUNK]->(d)
            SET r.index = row.chunk_index
            """,
            file_id=file_record.id,
            rows=chunk_rows,
        )

        next_rows = [
            {"a": chunks[index].id, "b": chunks[index + 1].id}
            for index in range(len(chunks) - 1)
        ]
        if next_rows:
            tx.run(
                """
                UNWIND $rows AS row
                MATCH (a:Document {id: row.a})
                MATCH (b:Document {id: row.b})
                MERGE (a)-[:NEXT_CHUNK]->(b)
                """,
                rows=next_rows,
            )

        if semantic_records and semantic_records.entities:
            tx.run(
                """
                UNWIND $rows AS row
                MERGE (e:Entity {id: row.id})
                SET e += row.props
                """,
                rows=[
                    {
                        "id": entity.id,
                        "props": {
                            "id": entity.id,
                            "usecase": entity.usecase,
                            "type": entity.type,
                            "name": entity.name,
                            "name_norm": entity.name_norm,
                            "aliases": entity.aliases,
                            "description": entity.description,
                            "updated_at": entity.updated_at,
                            "canonical_id": entity.canonical_id,
                            "embedding": entity.embedding,
                        },
                    }
                    for entity in semantic_records.entities
                ],
            )

        if semantic_records and semantic_records.mentions:
            tx.run(
                """
                UNWIND $rows AS row
                MATCH (d:Document {id: row.chunk_id})
                MATCH (e:Entity {id: row.entity_id})
                MERGE (d)-[r:MENTIONS]->(e)
                SET r.confidence = row.confidence,
                    r.updated_at = row.updated_at
                """,
                rows=[
                    {
                        "chunk_id": mention.chunk_id,
                        "entity_id": mention.entity_id,
                        "confidence": mention.confidence,
                        "updated_at": mention.updated_at,
                    }
                    for mention in semantic_records.mentions
                ],
            )

        if semantic_records and semantic_records.relations:
            tx.run(
                """
                UNWIND $rows AS row
                MATCH (a:Entity {id: row.source_entity_id})
                MATCH (b:Entity {id: row.target_entity_id})
                MERGE (a)-[r:REL {type: row.relation_type, source_chunk_id: row.source_chunk_id}]->(b)
                SET r.confidence = row.confidence,
                    r.updated_at = row.updated_at
                """,
                rows=[
                    {
                        "source_entity_id": relation.source_entity_id,
                        "target_entity_id": relation.target_entity_id,
                        "relation_type": relation.relation_type,
                        "source_chunk_id": relation.source_chunk_id,
                        "confidence": relation.confidence,
                        "updated_at": relation.updated_at,
                    }
                    for relation in semantic_records.relations
                ],
            )

        tx.run(
            """
            MATCH (e:Entity)
            WHERE e.usecase = $usecase
              AND NOT EXISTS { MATCH (:Document)-[:MENTIONS]->(e) }
            DETACH DELETE e
            """,
            usecase=file_record.usecase,
        )


def _extract_text_units(path: Path) -> list[TextUnit]:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return _extract_pdf_units(path)
    if suffix in {".md", ".txt"}:
        text = path.read_text(encoding="utf-8", errors="ignore")
        return [TextUnit(text=text, section=path.stem)] if text.strip() else []
    if suffix == ".xlsx":
        return _extract_xlsx_units(path)
    if suffix == ".docx":
        return _extract_docx_units(path)
    if suffix == ".ods":
        return _extract_ods_units(path)
    return []


def _extract_pdf_units(path: Path) -> list[TextUnit]:
    try:
        from pypdf import PdfReader
    except Exception as exc:  # pragma: no cover - dependency failure
        raise RuntimeError("pypdf is required for PDF ingestion") from exc

    reader = PdfReader(str(path))
    units: list[TextUnit] = []
    for index, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        if _is_blank_text(text):
            continue
        units.append(TextUnit(text=_normalize_text(text), page=index))
    return units


def _extract_xlsx_units(path: Path) -> list[TextUnit]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - dependency failure
        raise RuntimeError("openpyxl is required for XLSX ingestion") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    units: list[TextUnit] = []
    for sheet in workbook.worksheets:
        rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(value).strip() for value in row if value not in (None, "")]
            if cells:
                rows.append(" | ".join(cells))
        text = _normalize_text("\n".join(rows))
        if text:
            units.append(TextUnit(text=f"Sheet: {sheet.title}\n{text}", section=sheet.title))
    workbook.close()
    return units


def _extract_docx_units(path: Path) -> list[TextUnit]:
    with zipfile.ZipFile(path) as archive:
        xml_bytes = archive.read("word/document.xml")
    root = ET.fromstring(xml_bytes)
    paragraphs: list[str] = []
    namespace = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    for para in root.iter(f"{namespace}p"):
        text = _normalize_text("".join(para.itertext()))
        if text:
            paragraphs.append(text)
    text = "\n\n".join(paragraphs)
    return [TextUnit(text=text, section=path.stem)] if text else []


def _extract_ods_units(path: Path) -> list[TextUnit]:
    with zipfile.ZipFile(path) as archive:
        xml_bytes = archive.read("content.xml")
    root = ET.fromstring(xml_bytes)
    paragraphs: list[str] = []
    for elem in root.iter():
        if _local_name(elem.tag) not in {"p", "h"}:
            continue
        text = _normalize_text("".join(elem.itertext()))
        if text:
            paragraphs.append(text)
    text = "\n\n".join(paragraphs)
    return [TextUnit(text=text, section=path.stem)] if text else []


def _attach_embeddings(chunks: Sequence[DocumentChunk]) -> None:
    try:
        from backend.agents.memory.retriever import _get_embedding_model

        model = _get_embedding_model()
        if model is None:
            return
        texts = [chunk.text for chunk in chunks]
        embeddings = model.encode(texts)
        for chunk, embedding in zip(chunks, embeddings):
            object.__setattr__(chunk, "embedding", [float(value) for value in embedding.tolist()])
    except Exception as exc:
        logger.warning("Document embedding generation skipped: %s", exc)


def _attach_entity_embeddings(entities: Sequence[SemanticEntityRecord]) -> None:
    try:
        from backend.agents.memory.retriever import _get_embedding_model

        model = _get_embedding_model()
        if model is None or not entities:
            return
        texts = [f"{entity.type}: {entity.name}. {entity.description}".strip() for entity in entities]
        embeddings = model.encode(texts)
        for entity, embedding in zip(entities, embeddings):
            object.__setattr__(entity, "embedding", [float(value) for value in embedding.tolist()])
    except Exception as exc:
        logger.warning("Entity embedding generation skipped: %s", exc)


def _asset_tail(value: str) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    return re.split(r"[/:#]", raw)[-1]


def _entity_surface_keys(value: Any) -> set[str]:
    raw = str(value or "").strip()
    if not raw:
        return set()
    lowered = raw.lower()
    compact = re.sub(r"[^a-z0-9]+", "", lowered)
    return {key for key in {normalize_entity_name(raw), compact} if key}


def _run_coro_sync(coro: Any) -> Any:
    try:
        asyncio.get_running_loop()
    except RuntimeError:
        return asyncio.run(coro)

    with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
        return executor.submit(asyncio.run, coro).result()


async def _fetch_sindit_assets() -> list[dict[str, Any]]:
    from backend.agents.sindit.client import SinditClient

    async with SinditClient(base_url=SINDIT_API_URL, timeout=SINDIT_TIMEOUT_S) as client:
        # SINDIT requires a bearer token (Keycloak) and has no ``/kg/assets``
        # endpoint — assets are AbstractAsset nodes served via ``get_assets()``.
        # The old direct ``/kg/assets`` call authenticated nothing and 404'd, so
        # the canonical-id resolver never received any assets and document
        # entities were never linked to twin assets (canonical_id stayed null).
        await client.authenticate()
        assets = await client.get_assets()
    # Asset nodes key their identifier as ``uri``; the canonical-id resolver
    # reads ``iri``. Normalise so the resolver can index them.
    normalized: list[dict[str, Any]] = []
    for asset in assets or []:
        iri = str(asset.get("iri") or asset.get("uri") or "").strip()
        if not iri:
            continue
        normalized.append({**asset, "iri": iri})
    return normalized


def _build_document_entity_canonical_id_resolver(
    warnings: Optional[list[str]] = None,
) -> _DocumentEntityCanonicalIdResolver:
    assets: list[dict[str, Any]] = []
    if SINDIT_ENABLED:
        try:
            assets = list(_run_coro_sync(_fetch_sindit_assets()) or [])
        except Exception as exc:
            warning = f"SINDIT asset resolver unavailable: {exc}"
            logger.warning(warning)
            if warnings is not None:
                warnings.append(warning)
    return _DocumentEntityCanonicalIdResolver(assets)


def _entity_description_from_chunk(text: str, max_chars: int = 280) -> str:
    cleaned = _normalize_text(text)
    if len(cleaned) <= max_chars:
        return cleaned
    return cleaned[: max_chars - 3].rstrip() + "..."


def _chunk_text(text: str, *, chunk_chars: int, chunk_overlap: int) -> list[str]:
    normalized = _normalize_text(text)
    if not normalized:
        return []
    if len(normalized) <= chunk_chars:
        return [normalized]

    min_split = max(400, chunk_chars // 3)
    chunks: list[str] = []
    start = 0
    while start < len(normalized):
        end = min(start + chunk_chars, len(normalized))
        split = end
        if end < len(normalized):
            split = _best_split_point(normalized, start + min_split, end)
            if split <= start:
                split = end
        piece = normalized[start:split].strip()
        if piece:
            chunks.append(piece)
        if split >= len(normalized):
            break
        next_start = max(split - chunk_overlap, start + 1)
        if next_start <= start:
            next_start = split
        start = next_start
    return chunks


def _best_split_point(text: str, minimum: int, maximum: int) -> int:
    candidates = []
    for separator in ("\n\n", ". ", "\n", "; ", ", ", " "):
        index = text.rfind(separator, minimum, maximum)
        if index >= minimum:
            candidates.append(index + len(separator))
    return max(candidates) if candidates else maximum


def _usecase_for_path(path: Path, root: Path) -> str:
    relative = path.relative_to(root)
    top_level = relative.parts[0] if len(relative.parts) > 1 else None
    return resolve_usecase(
        usecase=top_level,
        source=relative.as_posix(),
        fallback_generic=True,
    ) or USECASE_GENERIC


def _subdirectory_for_path(path: Path, root: Path) -> Optional[str]:
    relative = path.relative_to(root)
    if len(relative.parts) <= 1:
        return None
    if _usecase_for_path(path, root) == USECASE_GENERIC:
        return "/".join(relative.parts[:-1]) or None
    return "/".join(relative.parts[1:-1]) or None


def _infer_document_type(path: Path) -> str:
    lowered = path.as_posix().lower()
    if "tool list" in lowered:
        return "tool_list"
    if any(token in lowered for token in ("operator", "functions", "documentacion", "documentación", "maquina", "machine")):
        return "manual"
    if any(token in lowered for token in ("guide", "guía", "guia", "systema integrado", "sistema integrado")):
        return "guide"
    if path.suffix.lower() in {".xlsx", ".ods"}:
        return "spreadsheet"
    if path.suffix.lower() == ".docx":
        return "docx"
    if path.suffix.lower() == ".md":
        return "reference"
    return "document"


def _infer_machine_ids(path: Path, usecase: str) -> list[str]:
    lowered = path.as_posix().lower()
    machine_ids: list[str] = []
    if "a1001" in lowered:
        machine_ids.append("MACHINE_A1")
    if "c1001" in lowered or "machine_c1" in lowered:
        machine_ids.append("c1001")
    if "b1001" in lowered:
        machine_ids.append("b1001")
    if "b1002" in lowered:
        machine_ids.append("b1002")

    if machine_ids:
        return list(dict.fromkeys(machine_ids))
    if usecase == "SITE_A":
        return ["MACHINE_A1"]
    if usecase == "SITE_C":
        return ["c1001"]
    if usecase == "SITE_B":
        return ["b1001", "b1002"]
    return []


def _machine_uri(machine_ids: Sequence[str]) -> Optional[str]:
    if len(machine_ids) != 1:
        return None
    slug = re.sub(r"[^a-z0-9_-]+", "-", machine_ids[0].strip().lower()).strip("-")
    if not slug:
        return None
    return f"urn:lfl:asset:{slug}"


def _guess_language(path: Path, sample: str) -> tuple[Optional[str], Optional[str]]:
    haystack = f"{path.name} {sample}".lower()
    scores = {
        "en": sum(haystack.count(token) for token in (" the ", " operator ", " functions ", " milling ", " guide ")),
        "es": sum(haystack.count(token) for token in (" de ", " la ", " máquina", " maquina", " guía", " guia", " documentación", "documentacion")),
        "de": sum(haystack.count(token) for token in (" und ", " der ", " die ", " das ", " site_c", "site_c")),
    }
    lang_code = max(scores, key=scores.get)
    if scores[lang_code] <= 0:
        return None, None
    lang_name = {"en": "English", "es": "Spanish", "de": "German"}[lang_code]
    return lang_code, lang_name


def _is_blank_text(text: str) -> bool:
    clean = _normalize_text(text).lower()
    if len(clean) < 20:
        return True
    return any(hint in clean for hint in _BLANK_PAGE_HINTS)


def _normalize_text(text: str) -> str:
    lines = [re.sub(r"\s+", " ", line).strip() for line in str(text or "").splitlines()]
    parts = [line for line in lines if line]
    return "\n".join(parts).strip()


def _stable_id(prefix: str, value: str) -> str:
    digest = hashlib.sha1(value.encode("utf-8")).hexdigest()
    return f"{prefix}:{digest}"


def _local_name(tag: str) -> str:
    return tag.split("}", 1)[-1]


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()
