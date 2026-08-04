"""Unified tool-master parser for the workbooks under ``data/tools/``.

The first implementation is intentionally conservative:

- Builder_b1 remains the primary Site_b geometry source.
- ``Site_b_Tool List Reviewed v2.xlsx`` only overlays values when the
  descriptions are compatible enough to trust the join by tool number.
- Site_a and Press_c are parsed directly from their workbook layouts.

The result is a pure in-process lookup surface that later SINDIT import and
runtime enrichment steps can consume without taking a graph dependency.
"""

from __future__ import annotations

import json
import logging
import re
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, Iterable, Optional

import openpyxl
import yaml
from pydantic import BaseModel

from .tool_dataset_decisions import resolve_confirmed_tool_context

logger = logging.getLogger(__name__)

REPO_ROOT = Path(__file__).resolve().parents[3]
TOOLS_DIR = REPO_ROOT / "data" / "tools"
MACHINE_FAMILIES_PATH = TOOLS_DIR / "machine_families.yaml"

FAMILY_BUILDER_B12 = "builder_b12"
FAMILY_MACHINE_A1 = "machine_a1"
FAMILY_PRESS_C_20_0482_010 = "press_c-20-0482-010"

# Per-dataset tool catalogs keyed by tool *id* (SG_Active_tool_name), not magazine
# slot. Used for datasets whose slot numbering does not match a shared master —
# Site_a_line2 shares much of the SITE_A magazine but its slot numbers differ, so its
# tools are resolved by id. The catalog holds only what is known for that dataset;
# unlisted tools fall back to the shared master (matched by id), then to the slot.
DATASET_TOOL_CATALOGS: dict[str, Path] = {
    "site_a_line2": REPO_ROOT / "data" / "Site_a_line2" / "site_a_line2_tools.yaml",
}

_OD_RE = re.compile(r"OD(\d{3,4})(?:\D|$)", re.IGNORECASE)
_L_RE = re.compile(r"L(\d{2,4})(?:\D|$)", re.IGNORECASE)
_DIAMETER_RE = re.compile(r"[Øø]\s*(\d+(?:[\.,]\d+)?)")
_MM_DIA_RE = re.compile(r"(\d+(?:[\.,]\d+)?)\s*MM\s*DIA", re.IGNORECASE)
_THREAD_RE = re.compile(r"^M\s*(\d+(?:[\.,]\d+)?)$", re.IGNORECASE)
_ARTICLE_RE = re.compile(r"Press_c_([0-9.]+)_", re.IGNORECASE)
_TOOL_NUMBER_RE = re.compile(r"^T?(\d+)$", re.IGNORECASE)


class ToolSpec(BaseModel):
    machine_family: str
    tool_number: int
    tool_id: str | None = None
    description: str | None = None
    tool_type: str | None = None
    diameter_mm: float | None = None
    teeth: int | None = None
    tool_length_mm: float | None = None
    insert_code: str | None = None
    tool_substrate: str | None = None
    source: str


def load_tool_master(*, refresh: bool = False) -> dict[tuple[str, int], ToolSpec]:
    """Return a copy of the merged tool-master dictionary."""
    if refresh:
        _load_tool_master_cached.cache_clear()
    master = _load_tool_master_cached()
    return {key: spec.model_copy(deep=True) for key, spec in master.items()}


def lookup(machine_family: str, tool_number: int | float | str | None) -> ToolSpec | None:
    """Return a copy of a parsed tool spec, or ``None`` when missing."""
    number = _canonical_tool_number(tool_number)
    if number is None:
        return None
    spec = _load_tool_master_cached().get((str(machine_family).strip().lower(), number))
    return spec.model_copy(deep=True) if spec is not None else None


def lookup_by_tool_id(machine_family: str | None, tool_id: str | None) -> ToolSpec | None:
    """Return a tool spec by tool *id* within a family (or any family), else ``None``."""
    tid = _text(tool_id)
    if tid is None:
        return None
    index = _master_by_tool_id_cached()
    fam = str(machine_family).strip().lower() if _text(machine_family) is not None else None
    spec = index.get((fam, tid.lower())) if fam else None
    if spec is None:
        spec = index.get((None, tid.lower()))  # cross-family fall back
    return spec.model_copy(deep=True) if spec is not None else None


def lookup_dataset_tool_catalog(dataset_id: str | None, tool_id: str | None) -> dict[str, Any] | None:
    """Return curated fields for a tool id from a dataset's tool catalog, else ``None``."""
    did = _text(dataset_id)
    tid = _text(tool_id)
    if did is None or tid is None:
        return None
    path = DATASET_TOOL_CATALOGS.get(did.strip().lower())
    if path is None:
        return None
    catalog = _load_dataset_tool_catalog_cached(str(path))
    return catalog.get(tid.lower())


@lru_cache(maxsize=8)
def _load_dataset_tool_catalog_cached(path_str: str) -> dict[str, dict[str, Any]]:
    path = Path(path_str)
    if not path.exists():
        return {}
    try:
        raw = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    except Exception:
        logger.exception("Failed to load dataset tool catalog %s", path)
        return {}
    tools = raw.get("tools") if isinstance(raw, dict) else None
    out: dict[str, dict[str, Any]] = {}
    if isinstance(tools, dict):
        for tid, fields in tools.items():
            key = _text(tid)
            if key is None or not isinstance(fields, dict):
                continue
            out[key.lower()] = {
                "tool_id": key,
                "tool_type": fields.get("tool_type"),
                "tool_diameter": _parse_float(fields.get("diameter_mm")),
                "tool_length": _parse_float(fields.get("tool_length_mm")),
                "tool_material": fields.get("tool_substrate") or fields.get("tool_material"),
                "num_teeth": _parse_int(fields.get("teeth")),
                "description": fields.get("description"),
            }
    return out


@lru_cache(maxsize=1)
def _master_by_tool_id_cached() -> dict[tuple[Optional[str], str], ToolSpec]:
    """Index the tool master by (family, tool_id_lower) and (None, tool_id_lower)."""
    index: dict[tuple[Optional[str], str], ToolSpec] = {}
    for spec in _load_tool_master_cached().values():
        tid = _text(spec.tool_id)
        if tid is None:
            continue
        index[(spec.machine_family.strip().lower(), tid.lower())] = spec
        index.setdefault((None, tid.lower()), spec)
    return index


def _spec_fields(spec: ToolSpec) -> dict[str, Any]:
    return {
        "tool_id": spec.tool_id,
        "tool_type": spec.tool_type,
        "tool_diameter": spec.diameter_mm,
        "tool_length": spec.tool_length_mm,
        "tool_material": spec.tool_substrate,
        "num_teeth": spec.teeth,
    }


def _slug(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", str(text).strip().lower()).strip("-") or "tool"


def resolve_tool_context(
    machine_family: str | None,
    tool_number: int | float | str | None,
    *,
    dataset_id: str | None = None,
    machine_id: str | None = None,
    raw_teeth: int | float | str | None = None,
    tool_id: str | None = None,
) -> dict[str, Any]:
    """Resolve tool-master-backed context fields for a dataset row or window.

    When *tool_id* (e.g. ``SG_Active_tool_name``) is supplied it is the authoritative
    tool identity: resolution prefers the dataset tool catalog, then the shared master
    matched by id, and does NOT fall back to the magazine slot (*tool_number*) — that
    avoids propagating wrong specs for datasets that renumber slots. When no tool id is
    supplied, resolution is by slot (confirmed decision, then master) as before.
    """
    context: dict[str, Any] = {}

    family_text = _text(machine_family)
    family = family_text.strip().lower() if family_text is not None else None
    if family:
        context["machine_family"] = family

    machine_text = _text(machine_id)
    if machine_text is not None:
        context["machine_id"] = machine_text

    number = _canonical_tool_number(tool_number)
    tool_id_text = _text(tool_id)
    if number is None and tool_id_text is None:
        return context

    # Resolve a fields dict, preferring the tool-id identity when available.
    fields: dict[str, Any] | None = None
    if tool_id_text is not None:
        fields = lookup_dataset_tool_catalog(dataset_id, tool_id_text)   # curated per-dataset
        if fields is None:
            spec = lookup_by_tool_id(family, tool_id_text)               # shared master, by id
            if spec is not None:
                fields = _spec_fields(spec)
    elif family and number is not None:
        confirmed = resolve_confirmed_tool_context(dataset_id, family, number) if dataset_id else None
        if confirmed is not None:
            fields = {
                "tool_id": confirmed.get("tool_id"),
                "tool_type": confirmed.get("tool_type"),
                "tool_diameter": confirmed.get("tool_diameter"),
                "tool_length": confirmed.get("tool_length"),
                "tool_material": confirmed.get("tool_material"),
                "num_teeth": _parse_int(confirmed.get("num_teeth")),
            }
        else:
            spec = lookup(family, number)
            if spec is not None:
                fields = _spec_fields(spec)

    # Identity: SG tool id (or resolved spec id) wins; else the legacy ``T{slot}``.
    if number is not None:
        context["tool_number"] = number
    authoritative_id = tool_id_text or (_text(fields.get("tool_id")) if fields else None)
    final_id = authoritative_id or (f"T{number}" if number is not None else None)
    if final_id is not None:
        context["tool_id"] = final_id
    if family:
        if tool_id_text is not None:
            context["sindit_tool_iri"] = f"urn:lfl:tool:{family}-{_slug(tool_id_text)}"
        elif number is not None:
            context["sindit_tool_iri"] = f"urn:lfl:tool:{family}-t{number}"

    if fields:
        if _text(fields.get("tool_type")) is not None:
            context["tool_type"] = fields["tool_type"]
        if fields.get("tool_diameter") is not None:
            context["tool_diameter"] = fields["tool_diameter"]
        if fields.get("tool_length") is not None:
            context["tool_length"] = fields["tool_length"]
        if _text(fields.get("tool_material")) is not None:
            context["tool_material"] = fields["tool_material"]

    teeth = fields.get("num_teeth") if fields else None
    if teeth is None:
        raw_teeth_value = _parse_int(raw_teeth)
        if raw_teeth_value is not None and raw_teeth_value > 0:
            teeth = raw_teeth_value
    if teeth is not None:
        context["num_teeth"] = int(teeth)

    return context


def load_machine_family_registry(
    path: Path | None = None,
    *,
    refresh: bool = False,
) -> dict[str, list[str]]:
    """Load the configured machine-family registry from YAML.

    The registry keys are canonical tool-master families (``builder_b12``,
    ``machine_a1``, ``press_c-20-0482-010``). Values are the case/machine
    identifiers observed by the loaders.
    """
    registry_path = Path(path) if path is not None else MACHINE_FAMILIES_PATH
    if refresh:
        _load_machine_family_registry_cached.cache_clear()
    cached = _load_machine_family_registry_cached(str(registry_path.resolve()))
    return {family: list(machine_ids) for family, machine_ids in cached.items()}


def resolve_machine_family(machine_id: str | None, path: Path | None = None) -> str:
    """Return the configured family for a machine/case id, or a slug fallback."""
    raw = _text(machine_id)
    if raw is None:
        return ""
    reverse_index = _build_machine_family_index(path)
    return reverse_index.get(_normalize_machine_key(raw), _slug_token(raw))


def dump_tool_master_json(path: Path | None = None) -> Path:
    """Write the parsed master to JSON for inspection and return the path."""
    target = path or TOOLS_DIR / "_master.json"
    target.parent.mkdir(parents=True, exist_ok=True)
    master = load_tool_master()
    payload = {
        f"{family}:{tool_number}": spec.model_dump(mode="json")
        for (family, tool_number), spec in sorted(master.items())
    }
    target.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")
    return target


@lru_cache(maxsize=1)
def _load_tool_master_cached() -> dict[tuple[str, int], ToolSpec]:
    master: dict[tuple[str, int], ToolSpec] = {}
    # Tool-master workbooks are site-supplied and optional: this distribution
    # ships no tool data. A missing or unreadable workbook degrades to "no tool
    # geometry known" rather than failing the request — enrichment is additive.
    for loader in (_load_builder_b12, _load_machine_a1, _load_press_c):
        try:
            chunk = loader()
        except FileNotFoundError:
            logger.debug("tool-master source unavailable: %s", loader.__name__)
            continue
        except Exception:
            logger.warning(
                "failed to read tool-master source %s — continuing without it",
                loader.__name__,
                exc_info=True,
            )
            continue
        for key, spec in chunk.items():
            existing = master.get(key)
            master[key] = _merge_specs(existing, spec) if existing is not None else spec
    if not master:
        logger.info(
            "No tool-master data found under %s — tool geometry enrichment disabled",
            TOOLS_DIR,
        )
    else:
        logger.info("Loaded %d tool-master entries from %s", len(master), TOOLS_DIR)
    return master


@lru_cache(maxsize=8)
def _load_machine_family_registry_cached(path_str: str) -> dict[str, tuple[str, ...]]:
    path = Path(path_str)
    if not path.is_file():
        logger.info("Machine-family registry %s not found; using empty mapping", path)
        return {}

    raw = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    families = raw.get("families", raw) if isinstance(raw, dict) else {}
    if not isinstance(families, dict):
        raise ValueError(f"Machine-family registry {path} must be a mapping")

    out: dict[str, tuple[str, ...]] = {}
    for family, payload in families.items():
        family_name = _text(family)
        if family_name is None:
            continue
        machine_ids: list[str] = []
        if isinstance(payload, dict):
            raw_machine_ids = payload.get("machine_ids", []) or []
        elif isinstance(payload, list):
            raw_machine_ids = payload
        else:
            raw_machine_ids = []
        for machine_id in raw_machine_ids:
            text = _text(machine_id)
            if text is not None:
                machine_ids.append(text)
        out[family_name.strip().lower()] = tuple(machine_ids)
    return out


def _build_machine_family_index(path: Path | None = None) -> dict[str, str]:
    registry = load_machine_family_registry(path)
    reverse: dict[str, str] = {}
    for family, machine_ids in registry.items():
        for machine_id in machine_ids:
            reverse[_normalize_machine_key(machine_id)] = family
    return reverse


def _load_machine_a1() -> dict[tuple[str, int], ToolSpec]:
    path = TOOLS_DIR / "site_a" / "Machine_a1.xlsx"
    rows = _iter_sheet_rows(path, header_row=1)
    out: dict[tuple[str, int], ToolSpec] = {}
    for row in rows:
        tool_number = _canonical_tool_number(row.get("T number"))
        if tool_number is None:
            continue
        tool_id = _text(row.get("ID"))
        desc1 = _text(row.get("Description 1"))
        desc2 = _text(row.get("Description 2"))
        spec = ToolSpec(
            machine_family=FAMILY_MACHINE_A1,
            tool_number=tool_number,
            tool_id=tool_id,
            description=desc2 or desc1,
            tool_type=_classify_tool_type(desc1, desc2, tool_id=tool_id),
            diameter_mm=_parse_site_a_diameter(desc1, desc2),
            tool_length_mm=_parse_site_a_length(desc1, desc2),
            source="site_a/Machine_a1.xlsx",
        )
        out[(spec.machine_family, spec.tool_number)] = spec
    return out


def _load_builder_b12() -> dict[tuple[str, int], ToolSpec]:
    path = TOOLS_DIR / "site_b" / "Builder_b1 2 Tooling Database.xlsx"
    rows = _iter_sheet_rows(path, header_row=1)
    out: dict[tuple[str, int], ToolSpec] = {}

    for row in rows:
        tool_number = _canonical_tool_number(row.get("Tool No"))
        if tool_number is None:
            continue
        description = _text(row.get("Description"))
        spec = ToolSpec(
            machine_family=FAMILY_BUILDER_B12,
            tool_number=tool_number,
            tool_id=f"T{tool_number:02d}",
            description=description,
            tool_type=_classify_tool_type(description),
            diameter_mm=_parse_diameter(description),
            teeth=_infer_builder_b1_teeth(description),
            tool_length_mm=_parse_float(row.get("TLO")),
            insert_code=_text(row.get("Builder_b1 2 inserts")),
            source="site_b/Builder_b1 2 Tooling Database.xlsx",
        )
        out[(spec.machine_family, spec.tool_number)] = spec

    for tool_number, reviewed in _load_reviewed_v2().items():
        key = (FAMILY_BUILDER_B12, tool_number)
        spec = out.get(key)
        if spec is None:
            out[key] = ToolSpec(
                machine_family=FAMILY_BUILDER_B12,
                tool_number=tool_number,
                tool_id=f"T{tool_number:02d}",
                description=reviewed["description"],
                tool_type=reviewed["tool_type"],
                diameter_mm=reviewed["diameter_mm"],
                teeth=reviewed["teeth"] or _infer_builder_b1_teeth(reviewed["description"]),
                source="site_b/Site_b_Tool List Reviewed v2.xlsx",
            )
            continue

        if _description_is_empty(spec.description):
            if reviewed["description"] is not None:
                spec.description = reviewed["description"]
            if spec.tool_type is None and reviewed["tool_type"] is not None:
                spec.tool_type = reviewed["tool_type"]
            if spec.diameter_mm is None and reviewed["diameter_mm"] is not None:
                spec.diameter_mm = reviewed["diameter_mm"]
            if spec.teeth is None:
                spec.teeth = reviewed["teeth"] or _infer_builder_b1_teeth(spec.description)
            spec.source = _append_source(spec.source, "site_b/Site_b_Tool List Reviewed v2.xlsx")
            continue

        if not _descriptions_compatible(spec.description, reviewed["description"]):
            continue

        if spec.diameter_mm is None and reviewed["diameter_mm"] is not None:
            spec.diameter_mm = reviewed["diameter_mm"]
        if spec.teeth is None and reviewed["teeth"] is not None:
            spec.teeth = reviewed["teeth"]
        if spec.teeth is None:
            spec.teeth = _infer_builder_b1_teeth(spec.description)
        spec.source = _append_source(spec.source, "site_b/Site_b_Tool List Reviewed v2.xlsx")

    return out


def _load_reviewed_v2() -> dict[int, Dict[str, Any]]:
    path = TOOLS_DIR / "site_b" / "Site_b_Tool List Reviewed v2.xlsx"
    rows = _iter_sheet_rows(path, header_row=8)
    out: dict[int, Dict[str, Any]] = {}
    for row in rows:
        tool_number = _canonical_tool_number(row.get("TOOL"))
        if tool_number is None:
            continue
        description = _text(row.get("DESCRIPTION"))
        out[tool_number] = {
            "description": description,
            "tool_type": _classify_tool_type(description),
            "diameter_mm": _parse_diameter(description),
            "teeth": _parse_int(row.get("NUMBER OF INSERTS")),
        }
    return out


def _load_press_c() -> dict[tuple[str, int], ToolSpec]:
    directory = TOOLS_DIR / "site_c"
    out: dict[tuple[str, int], ToolSpec] = {}
    for path in sorted(directory.glob("Press_c_*.xlsx")):
        if path.name.endswith("Zone.Identifier"):
            continue
        family = _press_c_family_from_name(path.name)
        rows = _iter_press_c_rows(path)
        for row in rows:
            tool_number = _canonical_tool_number(row.get("tool_number"))
            if tool_number is None:
                continue
            description = _text(row.get("description"))
            spec = ToolSpec(
                machine_family=family,
                tool_number=tool_number,
                description=description,
                tool_type=_classify_tool_type(description),
                diameter_mm=_parse_diameter(row.get("diameter")),
                tool_length_mm=_parse_float(row.get("length")),
                source=f"site_c/{path.name}",
            )
            key = (spec.machine_family, spec.tool_number)
            existing = out.get(key)
            out[key] = _merge_specs(existing, spec) if existing is not None else spec
    return out


def _iter_sheet_rows(path: Path, *, header_row: int) -> Iterable[Dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [str(value).strip() if value is not None else "" for value in next(
        ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
    )]

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        payload = {
            headers[index]: value
            for index, value in enumerate(row)
            if index < len(headers) and headers[index]
        }
        if any(_text(value) is not None for value in payload.values()):
            yield payload


def _iter_press_c_rows(path: Path) -> Iterable[Dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    description_memory: str | None = None
    for row in ws.iter_rows(min_row=10, values_only=True):
        values = list(row)
        tool_number = values[0] if len(values) > 0 else None
        description = values[2] if len(values) > 2 else None
        diameter = values[11] if len(values) > 11 else None
        length = values[14] if len(values) > 14 else None
        note = values[17] if len(values) > 17 else None

        if _canonical_tool_number(tool_number) is None:
            continue

        description_text = _text(description)
        if description_text is not None:
            ditto_suffix = _press_c_ditto_suffix(description_text)
            if ditto_suffix is not None:
                description_text = " ".join(part for part in (description_memory, ditto_suffix) if part) or description_memory
        if description_text in {None, '"     -     "'}:
            description_text = description_memory
        elif description_text is not None:
            description_memory = description_text

        yield {
            "tool_number": tool_number,
            "description": description_text,
            "diameter": diameter,
            "length": length,
            "note": note,
        }


def _press_c_family_from_name(filename: str) -> str:
    match = _ARTICLE_RE.search(filename)
    if not match:
        return FAMILY_PRESS_C_20_0482_010
    article = match.group(1)
    return f"press_c-{article.replace('.', '-')}".lower()


def _canonical_tool_number(value: Any) -> int | None:
    text = _text(value)
    if text is None:
        return None
    match = _TOOL_NUMBER_RE.match(text)
    if match:
        return int(match.group(1))
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def _text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_machine_key(value: str) -> str:
    return " ".join(str(value).strip().lower().split())


def _slug_token(value: str) -> str:
    out = []
    for ch in str(value or "").strip():
        if ch.isalnum() or ch in ("-", "_"):
            out.append(ch.lower())
        elif ch.isspace() or ch in (":", "/", "\\", ".", ","):
            out.append("-")
    slug = "".join(out).strip("-")
    return slug or "unknown"


def _parse_int(value: Any) -> int | None:
    text = _text(value)
    if text is None:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _parse_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _text(value)
    if text is None:
        return None
    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _parse_site_a_diameter(*texts: str | None) -> float | None:
    for text in texts:
        if not text:
            continue
        match = _OD_RE.search(text)
        if match:
            return float(match.group(1)) / 10.0
    for text in texts:
        diameter = _parse_diameter(text)
        if diameter is not None:
            return diameter
    return None


def _parse_site_a_length(*texts: str | None) -> float | None:
    for text in texts:
        if not text:
            continue
        match = _L_RE.search(text)
        if match:
            return float(match.group(1))
    return None


def _parse_diameter(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _text(value)
    if text is None:
        return None

    for regex in (_DIAMETER_RE, _MM_DIA_RE):
        match = regex.search(text)
        if match:
            return float(match.group(1).replace(",", "."))

    match = _THREAD_RE.match(text)
    if match:
        return float(match.group(1).replace(",", "."))

    numeric = _parse_float(text)
    if numeric is not None:
        return numeric
    return None


def _classify_tool_type(*texts: str | None, tool_id: str | None = None) -> str | None:
    haystack = " ".join(text.lower() for text in texts if text)
    tool_id_text = (tool_id or "").upper()

    if "probe" in haystack or "palpadora" in haystack or tool_id_text.startswith("S"):
        return "probe"
    if any(token in haystack for token in ("gewindebohrer", "gew.bohrer", "gewinde-bohrer", "gew.former", "gewinde-former", " ardatza ")) or " tap" in haystack or haystack.startswith("tap") or haystack.startswith("ardatza"):
        return "tap"
    if "senker" in haystack or "countersink" in haystack:
        return "countersink"
    if "bore" in haystack or "boring" in haystack:
        return "bore"
    if "bohrer" in haystack or "drill" in haystack or tool_id_text.startswith("D"):
        return "drill"
    if any(token in haystack for token in ("mill", "milling", "cutter", "fräser", "fresa", "messerkopf")):
        return "mill"
    if tool_id_text.startswith("M"):
        return "mill"
    return None


def _press_c_ditto_suffix(text: str) -> str | None:
    cleaned = str(text or "").strip()
    if not cleaned.startswith('"'):
        return None
    suffix = cleaned.replace('"', " ").replace("-", " ")
    suffix = " ".join(suffix.split())
    return suffix or None


def _descriptions_compatible(primary: str | None, secondary: str | None) -> bool:
    primary_text = (primary or "").strip()
    secondary_text = (secondary or "").strip()
    if not primary_text or not secondary_text:
        return False
    if primary_text.upper() == "EMPTY":
        return False

    primary_type = _classify_tool_type(primary_text)
    secondary_type = _classify_tool_type(secondary_text)
    if primary_type and secondary_type and primary_type != secondary_type:
        return False

    primary_diameter = _parse_diameter(primary_text)
    secondary_diameter = _parse_diameter(secondary_text)
    if (
        primary_diameter is not None
        and secondary_diameter is not None
        and abs(primary_diameter - secondary_diameter) > 2.0
    ):
        return False

    primary_tokens = _meaningful_tokens(primary_text)
    secondary_tokens = _meaningful_tokens(secondary_text)
    return bool(primary_tokens & secondary_tokens) or (
        primary_type is not None and primary_type == secondary_type
    )


def _description_is_empty(description: str | None) -> bool:
    text = (description or "").strip()
    return not text or text.upper() == "EMPTY"


def _infer_builder_b1_teeth(description: str | None) -> int | None:
    text = (description or "").strip().upper()
    if not text:
        return None
    if "FINISH BORE" in text:
        return 1
    if "ROUGH BORE" in text:
        return 2
    return None


def _meaningful_tokens(text: str) -> set[str]:
    stop_words = {
        "dia",
        "mm",
        "finish",
        "rough",
        "solid",
        "carbide",
        "face",
        "long",
        "short",
    }
    return {
        token
        for token in re.findall(r"[a-z]+", text.lower())
        if len(token) > 2 and token not in stop_words
    }


def _merge_specs(existing: ToolSpec | None, incoming: ToolSpec) -> ToolSpec:
    if existing is None:
        return incoming
    if existing.tool_id is None and incoming.tool_id is not None:
        existing.tool_id = incoming.tool_id
    if existing.description is None and incoming.description is not None:
        existing.description = incoming.description
    if existing.tool_type is None and incoming.tool_type is not None:
        existing.tool_type = incoming.tool_type
    if existing.diameter_mm is None and incoming.diameter_mm is not None:
        existing.diameter_mm = incoming.diameter_mm
    if existing.teeth is None and incoming.teeth is not None:
        existing.teeth = incoming.teeth
    if existing.tool_length_mm is None and incoming.tool_length_mm is not None:
        existing.tool_length_mm = incoming.tool_length_mm
    if existing.insert_code is None and incoming.insert_code is not None:
        existing.insert_code = incoming.insert_code
    if existing.tool_substrate is None and incoming.tool_substrate is not None:
        existing.tool_substrate = incoming.tool_substrate
    existing.source = _append_source(existing.source, incoming.source)
    return existing


def _append_source(current: str, new_source: str) -> str:
    if not current:
        return new_source
    existing = {part.strip() for part in current.split(" + ") if part.strip()}
    if new_source in existing:
        return current
    return f"{current} + {new_source}"


__all__ = [
    "FAMILY_MACHINE_A1",
    "FAMILY_BUILDER_B12",
    "FAMILY_PRESS_C_20_0482_010",
    "MACHINE_FAMILIES_PATH",
    "ToolSpec",
    "dump_tool_master_json",
    "load_tool_master",
    "load_machine_family_registry",
    "lookup",
    "resolve_tool_context",
    "resolve_machine_family",
]