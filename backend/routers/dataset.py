"""Dataset Explorer router — merged timeseries browser.

Serves preprocessed merged CSVs from one or more recording sessions with
LTTB downsampling and process-annotation regions derived from OF workbook
data (PGM LINE → Cnc_ProgramActive_BlockNumber_RT mapping).
"""

from __future__ import annotations

import json
import logging
import math
import os
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import numpy as np
import pandas as pd
from fastapi import APIRouter, Query
from backend.json_utils import finite_float as _safe_float

import re

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/dataset", tags=["dataset"])

# ── Paths ────────────────────────────────────────────────────────────────────

_DATA_DIR = Path(os.environ.get("DATASET_DATA_DIR", "data/Site_a_line2"))
_MERGED_DIR = _DATA_DIR / "merged"
_OFS_XLSX = _DATA_DIR / os.environ.get("DATASET_OFS_TRANSLATED", "OFs_translated.xlsx")
_OFS_XLSX_ORIG = _DATA_DIR / os.environ.get("DATASET_OFS_ORIGINAL", "OFs.xlsx")


@lru_cache(maxsize=1)
def _get_sessions_registry() -> Dict[str, Dict[str, str]]:
    """Return session registry from env JSON or merged CSV discovery.

    Env override accepts JSON string in ``DATASET_SESSIONS_JSON``:
    {
      "session1": {"file": "foo.csv", "label": "Session 1", "date_range": "", "channels": ""}
    }
    """
    raw = os.environ.get("DATASET_SESSIONS_JSON", "").strip()
    if raw:
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                return {
                    str(k): {str(kk): str(vv) for kk, vv in dict(v).items()}
                    for k, v in parsed.items()
                    if isinstance(v, dict) and "file" in v
                }
        except Exception:
            logger.warning("Invalid DATASET_SESSIONS_JSON; falling back to auto-discovery")

    csvs = sorted(_MERGED_DIR.glob("*.csv"))
    sessions: Dict[str, Dict[str, str]] = {}
    for i, p in enumerate(csvs, start=1):
        sid = f"session{i}"
        sessions[sid] = {
            "file": p.name,
            "label": f"Session {i}",
            "date_range": "",
            "channels": "",
        }
    return sessions

# ── Curated defaults ─────────────────────────────────────────────────────────

_DEFAULT_CHANNELS = [
    "Spindle_Power_percent",
    "SpindleSpeedActual",
    "Axis_FeedRate_actual",
    "Monit_chatter_detection_Chatter_amplitude_mm_s_acc1",
]

# ── Helpers ──────────────────────────────────────────────────────────────────
def _jsonable(v: Any) -> Any:
    """Convert numpy/openpyxl values to JSON-safe Python natives."""
    if v is None:
        return None
    if isinstance(v, (np.integer,)):
        return int(v)
    if isinstance(v, (np.floating,)):
        f = float(v)
        return None if (math.isnan(f) or math.isinf(f)) else f
    if isinstance(v, np.bool_):
        return bool(v)
    if isinstance(v, np.ndarray):
        return v.tolist()
    if isinstance(v, dict):
        return {k: _jsonable(val) for k, val in v.items()}
    if isinstance(v, (list, tuple)):
        return [_jsonable(item) for item in v]
    return v


def _lttb_downsample(x: np.ndarray, y: np.ndarray, target: int) -> Tuple[np.ndarray, np.ndarray]:
    """Largest-Triangle-Three-Buckets downsampling."""
    n = len(x)
    if n <= target:
        return x, y
    bucket_size = (n - 2) / (target - 2)
    out_x = [x[0]]
    out_y = [y[0]]
    a_idx = 0
    for i in range(1, target - 1):
        b_start = int(i * bucket_size) + 1
        b_end = min(int((i + 1) * bucket_size) + 1, n)
        c_start = int((i + 1) * bucket_size) + 1
        c_end = min(int((i + 2) * bucket_size) + 1, n)
        avg_x = np.mean(x[c_start:c_end]) if c_start < c_end else x[-1]
        avg_y = np.mean(y[c_start:c_end]) if c_start < c_end else y[-1]
        max_area = -1.0
        best = b_start
        for j in range(b_start, b_end):
            area = abs(
                (x[a_idx] - avg_x) * (y[j] - out_y[-1])
                - (x[a_idx] - x[j]) * (avg_y - out_y[-1])
            )
            if area > max_area:
                max_area = area
                best = j
        out_x.append(x[best])
        out_y.append(y[best])
        a_idx = best
    out_x.append(x[-1])
    out_y.append(y[-1])
    return np.array(out_x), np.array(out_y)


# ── Cached loaders ───────────────────────────────────────────────────────────

@lru_cache(maxsize=1)
def _load_all_sessions() -> Dict[str, pd.DataFrame]:
    """Load all session CSVs into memory, keyed by session id."""
    result: Dict[str, pd.DataFrame] = {}
    sessions = _get_sessions_registry()
    for sess_id, info in sessions.items():
        csv_path = _MERGED_DIR / info["file"]
        if not csv_path.exists():
            logger.warning("Session CSV not found: %s", csv_path)
            continue
        df = pd.read_csv(csv_path, low_memory=False)
        if "Date" in df.columns:
            df["Date"] = pd.to_datetime(df["Date"], errors="coerce", utc=True)
            df.sort_values("Date", inplace=True)
            df.reset_index(drop=True, inplace=True)
        df["_session"] = sess_id
        result[sess_id] = df
        logger.info("Loaded %s: %d rows × %d cols", sess_id, len(df), len(df.columns))
    if not result:
        raise FileNotFoundError("No session CSVs found in " + str(_MERGED_DIR))
    return result


def _load_merged_df(session: str = "") -> pd.DataFrame:
    """Return merged DataFrame for a specific session, or all sessions combined."""
    all_sessions = _load_all_sessions()
    if session and session in all_sessions:
        return all_sessions[session]
    # Combine all sessions — use only common columns
    frames = list(all_sessions.values())
    if len(frames) == 1:
        return frames[0]
    common_cols = set(frames[0].columns)
    for f in frames[1:]:
        common_cols &= set(f.columns)
    common_cols = sorted(common_cols)
    return pd.concat([f[common_cols] for f in frames], ignore_index=True)


@lru_cache(maxsize=1)
def _load_milestones() -> Dict[str, Any]:
    """Parse OFs xlsx into structured dicts, using both translated and
    untranslated files.  The key insight is that columns J/K (A-side) or
    I/J (B-side) contain the **actual CNC block numbers** ("Variable Value
    Prev/Next PGM Line") that correspond to
    ``Cnc_ProgramActive_BlockNumber_RT`` in the merged CSV.  Column F
    ("PGM LINE") is the program *line* number in the NC source — a
    different coordinate system.
    """

    # Prefer explicit translated/original paths, then any OF*.xlsx fallback.
    xlsx_path = _OFS_XLSX if _OFS_XLSX.exists() else _OFS_XLSX_ORIG
    if not xlsx_path.exists():
        candidates = sorted(_DATA_DIR.glob("*OF*.xlsx"))
        if candidates:
            xlsx_path = candidates[0]
    if not xlsx_path.exists():
        raise FileNotFoundError(f"OFs xlsx not found under {_DATA_DIR}")

    import openpyxl

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    # Also load untranslated as supplementary if available
    wb_orig = None
    if _OFS_XLSX_ORIG.exists() and xlsx_path != _OFS_XLSX_ORIG:
        try:
            wb_orig = openpyxl.load_workbook(str(_OFS_XLSX_ORIG), data_only=True)
        except Exception:
            pass

    result: Dict[str, Any] = {"sheets": {}}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        side_upper = sheet_name.upper()
        side = "A" if ("A-SIDE" in side_upper or "AD2618124" in side_upper) \
            else "B" if ("B-SIDE" in side_upper or "BD2618124" in side_upper) \
            else sheet_name

        # ── Detect header row ─────────────────────────────────────────────
        header_row = 1
        _HEADER_MARKERS = {"PGM LINE", "L PGM", "TOOL ID", "TOOL NAME"}
        for r in range(1, min(ws.max_row + 1, 15)):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v and str(v).upper().strip() in _HEADER_MARKERS:
                    header_row = r
                    break
            if header_row != 1:
                break

        # ── Build column map from header names ────────────────────────────
        # Handles both translated (English) and untranslated (Basque/Spanish)
        # headers, and the B-side column shift.
        col_map: Dict[str, int] = {}
        of_columns: List[Dict[str, Any]] = []

        for c in range(1, ws.max_column + 1):
            h = ws.cell(header_row, c).value
            if h is None:
                continue
            h_str = str(h).strip()
            h_upper = h_str.upper()

            if h_upper in ("PGM LINE", "L PGM"):
                col_map["pgm_line"] = c
            elif h_upper in ("OPERATION", "OPERAZIOA"):
                col_map["operation"] = c
            elif h_upper in ("HEAD", "KABEZALA"):
                col_map["head"] = c
            elif h_upper == "TOOL ID":
                col_map["tool_id"] = c
            elif h_upper == "TOOL NAME":
                col_map["tool_name"] = c
            elif h_upper in ("PROCESS", "PROZESUA"):
                col_map["process"] = c
            elif h_upper in ("SUB-PROCESS", "AZPIPROZESUA"):
                col_map["sub_process"] = c
            elif any(kw in h_upper for kw in ("PREV", "ANTERIOR")):
                col_map["block_prev"] = c
            elif any(kw in h_upper for kw in ("NEXT", "POSTERIOR")):
                col_map["block_next"] = c
            elif h_upper.startswith("OF"):
                of_num = h_str.replace("OF", "").replace(" ", "").strip()
                of_columns.append({"header": h_str, "of_number": of_num, "col": c})

        # ── Read data rows ────────────────────────────────────────────────
        # Include rows even without PGM LINE if they carry useful data
        # (block_prev/block_next, process names, or OF annotations).
        rows: List[Dict[str, Any]] = []
        pgm_col = col_map.get("pgm_line", 6)

        for r in range(header_row + 1, ws.max_row + 1):
            pgm_val = ws.cell(r, pgm_col).value
            pgm_line: Optional[int] = None
            if pgm_val is not None:
                try:
                    pgm_line = int(float(pgm_val))
                except (TypeError, ValueError):
                    pass

            row_data: Dict[str, Any] = {}
            if pgm_line is not None:
                row_data["pgm_line"] = pgm_line

            for key, c in col_map.items():
                if key == "pgm_line":
                    continue
                v = ws.cell(r, c).value
                if v is not None:
                    if key in ("block_prev", "block_next"):
                        try:
                            row_data[key] = int(float(v))
                        except (TypeError, ValueError):
                            pass
                    else:
                        row_data[key] = str(v).strip() if isinstance(v, str) else v

            # OF-specific annotations
            of_data: Dict[str, Any] = {}
            for of_col in of_columns:
                v = ws.cell(r, of_col["col"]).value
                if v is not None:
                    of_data[of_col["of_number"]] = str(v).strip() if isinstance(v, str) else v
            if of_data:
                row_data["of_annotations"] = of_data

            # Skip truly empty rows (no PGM LINE, no block data, no process,
            # no OF annotations).
            has_blocks = "block_prev" in row_data or "block_next" in row_data
            has_process = bool(row_data.get("process"))
            has_of = bool(of_data)
            if pgm_line is None and not has_blocks and not has_process and not has_of:
                continue

            rows.append(row_data)

        # ── Cross-reference with untranslated file to fill gaps ───────────
        if wb_orig and rows:
            _cross_reference_original(wb_orig, side, rows)

        # For rows missing block_next, infer from the next row's block_prev
        for i in range(len(rows) - 1):
            if "block_next" not in rows[i] and "block_prev" in rows[i + 1]:
                rows[i]["block_next"] = rows[i + 1]["block_prev"]

        # ── Summaries ─────────────────────────────────────────────────────
        of_ids = [of_col["of_number"] for of_col in of_columns]

        breakage_summary: Dict[str, List[Dict[str, Any]]] = {}
        for row in rows:
            anns = row.get("of_annotations", {})
            for oid, val in anns.items():
                if isinstance(val, str) and any(
                    kw in val.lower()
                    for kw in ("broken", "chipped", "breakage", "rotura")
                ):
                    breakage_summary.setdefault(oid, []).append({
                        "pgm_line": row["pgm_line"],
                        "process": row.get("process"),
                        "annotation": val,
                    })

        tool_summary: List[Dict[str, Any]] = []
        seen_tools: set = set()
        for row in rows:
            tid = row.get("tool_id")
            tname = row.get("tool_name")
            if tid and tid not in seen_tools:
                seen_tools.add(tid)
                tool_summary.append({
                    "tool_id": tid,
                    "tool_name": tname,
                    "first_pgm_line": row.get("pgm_line"),
                })

        result["sheets"][side] = {
            "name": sheet_name,
            "rows": rows,
            "of_columns": of_columns,
            "of_ids": of_ids,
            "breakage_summary": breakage_summary,
            "tool_summary": tool_summary,
            "total_milestones": len(rows),
        }

    return result


def _cross_reference_original(
    wb_orig,
    side: str,
    rows: List[Dict[str, Any]],
) -> None:
    """Fill missing fields from the untranslated xlsx (same structure,
    possibly more complete data in places)."""

    # Find matching sheet
    ws_orig = None
    for sn in wb_orig.sheetnames:
        su = sn.upper()
        if side == "A" and "AD2618124" in su:
            ws_orig = wb_orig[sn]
        elif side == "B" and "BD2618124" in su:
            ws_orig = wb_orig[sn]
    if ws_orig is None:
        return

    # Detect header row in original
    header_row = 1
    for r in range(1, min(ws_orig.max_row + 1, 15)):
        for c in range(1, ws_orig.max_column + 1):
            v = ws_orig.cell(r, c).value
            if v and str(v).upper().strip() in ("L PGM", "TOOL ID"):
                header_row = r
                break
        if header_row != 1:
            break

    # Build column map for original
    col_map_orig: Dict[str, int] = {}
    for c in range(1, ws_orig.max_column + 1):
        h = ws_orig.cell(header_row, c).value
        if h is None:
            continue
        h_upper = str(h).upper().strip()
        if h_upper == "L PGM":
            col_map_orig["pgm_line"] = c
        elif h_upper == "PROZESUA":
            col_map_orig["process"] = c
        elif h_upper == "AZPIPROZESUA":
            col_map_orig["sub_process"] = c
        elif "ANTERIOR" in h_upper:
            col_map_orig["block_prev"] = c
        elif "POSTERIOR" in h_upper:
            col_map_orig["block_next"] = c

    if "pgm_line" not in col_map_orig:
        return

    # Build pgm_line → original row data
    orig_by_pgm: Dict[int, Dict[str, Any]] = {}
    for r in range(header_row + 1, ws_orig.max_row + 1):
        pv = ws_orig.cell(r, col_map_orig["pgm_line"]).value
        if pv is None:
            continue
        try:
            pgm = int(float(pv))
        except (TypeError, ValueError):
            continue
        orig_row: Dict[str, Any] = {}
        for key, c in col_map_orig.items():
            if key == "pgm_line":
                continue
            v = ws_orig.cell(r, c).value
            if v is not None:
                if key in ("block_prev", "block_next"):
                    try:
                        orig_row[key] = int(float(v))
                    except (TypeError, ValueError):
                        pass
                else:
                    orig_row[key] = str(v).strip() if isinstance(v, str) else v
        orig_by_pgm[pgm] = orig_row

    # Fill gaps in translated rows
    for row in rows:
        pgm = row.get("pgm_line")
        if pgm is None:
            continue
        orig = orig_by_pgm.get(pgm)
        if not orig:
            continue
        for key in ("block_prev", "block_next", "process", "sub_process"):
            if key not in row and key in orig:
                row[key] = orig[key]


def _build_process_regions(
    df_of: pd.DataFrame,
    milestones: Dict[str, Any],
    of_id: str,
    side: str,
) -> List[Dict[str, Any]]:
    """Map milestone transitions to CSV timestamps.

    The CNC program (e.g. BD2618124ESP.H) machines **both** sides of the
    part in a single run.  A-side milestones occupy block numbers 1000-2560
    and B-side milestones occupy 5000-6630.  Between milestone transitions
    the CNC executes sub-programs (LBL calls) whose block numbers are in
    a completely different range.

    Strategy: merge milestones from **both** A and B sheets, build a
    combined block → milestone lookup, then scan the CSV chronologically.
    When the block number enters a milestone's [block_prev, block_next]
    range, mark it as the active milestone.  All time from that point
    until the next milestone detection belongs to that operation.
    """
    block_col = "Cnc_ProgramActive_BlockNumber_RT"
    if block_col not in df_of.columns or "_seconds" not in df_of.columns:
        return []

    blocks = df_of[block_col].values.astype(float)
    timestamps_s = df_of["_seconds"].values.astype(float)

    # ── Merge milestones from BOTH sides ─────────────────────────────────
    all_ms: List[Dict[str, Any]] = []
    for side_key in ("A", "B"):
        sheet = milestones["sheets"].get(side_key)
        if not sheet:
            continue
        for mrow in sheet["rows"]:
            if "block_prev" in mrow:
                enriched = dict(mrow)
                enriched["_side"] = side_key
                all_ms.append(enriched)

    if not all_ms:
        return []

    all_ms.sort(key=lambda m: m["block_prev"])

    # ── Propagate process names forward ──────────────────────────────
    # When a milestone has no process label, use fallbacks:
    # operation → sub_process → tool_name → previous milestone's process.
    last_process = ""
    for mrow in all_ms:
        proc = mrow.get("process")
        if proc and str(proc).strip():
            last_process = str(proc).strip()
        else:
            # Try fallback fields
            fallback = (
                mrow.get("operation")
                or mrow.get("sub_process")
                or mrow.get("tool_name")
            )
            if fallback and str(fallback).strip():
                mrow["process"] = str(fallback).strip()
            elif last_process:
                mrow["process"] = last_process

    # Build a dict mapping each block number in every milestone's
    # transition window [block_prev, block_next] → index in all_ms.
    block_to_ms_idx: Dict[int, int] = {}
    for i, mrow in enumerate(all_ms):
        bp = int(mrow["block_prev"])
        bn = int(mrow.get("block_next", bp + 5))
        for b in range(bp, bn + 1):
            block_to_ms_idx[b] = i

    # ── Chronological scan ───────────────────────────────────────────────
    step = max(1, len(blocks) // 8000)  # finer sampling for accuracy
    sampled_idx = list(range(0, len(blocks), step))
    if sampled_idx[-1] != len(blocks) - 1:
        sampled_idx.append(len(blocks) - 1)

    regions: List[Dict[str, Any]] = []
    current_ms_idx: Optional[int] = None
    region_start_s: float = 0.0

    def _flush_region(end_s: float) -> None:
        nonlocal current_ms_idx
        if current_ms_idx is None:
            return
        if end_s - region_start_s < 0.5:
            return  # skip sub-second noise
        mrow = all_ms[current_ms_idx]
        of_annotation = mrow.get("of_annotations", {}).get(of_id)
        is_breakage = False
        if isinstance(of_annotation, str) and any(
            kw in of_annotation.lower()
            for kw in ("broken", "chipped", "breakage", "rotura")
        ):
            is_breakage = True
        ms_side = mrow.get("_side", side)
        regions.append({
            "start_s": region_start_s,
            "end_s": round(end_s, 1),
            "label": str(mrow.get("process") or "(unnamed)"),
            "sub_process": str(mrow.get("sub_process") or ""),
            "tool_id": _jsonable(mrow.get("tool_id")),
            "tool_name": str(mrow.get("tool_name") or ""),
            "pgm_line": _jsonable(mrow.get("pgm_line")),
            "of_id": of_id,
            "of_annotation": str(of_annotation) if of_annotation else None,
            "is_breakage": bool(is_breakage),
            "severity": "critical" if is_breakage else "info",
            "block_prev": int(mrow["block_prev"]),
            "block_next": int(mrow.get("block_next", mrow["block_prev"] + 5)),
            "side": ms_side,
        })

    for si in sampled_idx:
        bval = blocks[si]
        if not np.isfinite(bval):
            continue
        bint = int(bval)
        ts = float(timestamps_s[si])

        # Check if this block is in a milestone transition window
        ms_idx = block_to_ms_idx.get(bint)

        if ms_idx is not None and ms_idx != current_ms_idx:
            # New milestone detected — flush previous region
            _flush_region(ts)
            current_ms_idx = ms_idx
            region_start_s = round(ts, 1)

    # Close last region
    if current_ms_idx is not None:
        _flush_region(float(timestamps_s[sampled_idx[-1]]))

    return regions


# Picture-reference pattern:  P0, P1, P7a, P7b, P9 etc.
_PIC_RE = re.compile(r'\bP(\d+[a-z]?)\b')

# Keywords indicating condition categories
_BREAKAGE_KW = ("broken", "chipped", "breakage", "rotura", "halted")
_WEAR_KW = ("wear", "worn")
_OK_KW = ("ok", "no evident wear", "normal", "new inserts")


def _extract_annotations(
    regions: List[Dict[str, Any]],
    milestones_data: Dict[str, Any],
    of_id: str,
) -> List[Dict[str, Any]]:
    """Extract ALL annotations for a specific OF from milestone data.

    Pulls annotation text directly from milestone rows (not only from
    regions), so annotations are not lost when a milestone's block range
    is too short to produce a visible region.

    Uses region end times for positioning; falls back to linear
    interpolation across known region times when no matching region exists.
    """
    if not of_id:
        return []

    # Build pgm_line → region mapping for timestamp lookup
    pgm_to_region: Dict[Any, Dict[str, Any]] = {}
    block_to_region: Dict[int, Dict[str, Any]] = {}
    for r in regions:
        pgm = r.get("pgm_line")
        if pgm is not None:
            pgm_to_region[pgm] = r
        bp = r.get("block_prev")
        if bp is not None:
            block_to_region[int(bp)] = r

    # Ordered (pgm_line, time) pairs for interpolation fallback
    timed_pgms = sorted(
        [(r["pgm_line"], r["end_s"]) for r in regions if r.get("pgm_line") is not None],
        key=lambda x: x[0],
    )

    seen: Set[Tuple[Any, str]] = set()
    annotations: List[Dict[str, Any]] = []

    # Only A-side has OF columns; B-side has none
    for side_key in ("A",):
        sheet = milestones_data.get("sheets", {}).get(side_key)
        if not sheet:
            continue

        for mrow in sheet["rows"]:
            of_anns = mrow.get("of_annotations", {})
            text = of_anns.get(of_id)
            if not text or not isinstance(text, str):
                continue
            text = text.strip()
            if not text:
                continue

            # Skip pure OF-header rows (e.g. "OF 100003890")
            if text.upper().startswith("OF") and text.replace(" ", "")[2:].isdigit():
                continue
            # Skip date-only rows (e.g. "04/03 - 15:20")
            if re.match(r'^\d{2}/\d{2}\s*-\s*\d', text):
                continue

            pgm = mrow.get("pgm_line")
            block_prev = mrow.get("block_prev")
            dedup_key = (pgm or block_prev, text)
            if dedup_key in seen:
                continue
            seen.add(dedup_key)

            # Get time from matching region (by pgm_line or block_prev),
            # or interpolate
            region = None
            if pgm is not None:
                region = pgm_to_region.get(pgm)
            if region is None and block_prev is not None:
                region = block_to_region.get(int(block_prev))

            if region:
                time_s = region["end_s"]
            else:
                time_s = _interpolate_time(pgm, timed_pgms)
                if time_s is None and block_prev is not None:
                    # Try interpolation with block_prev as proxy
                    time_s = _interpolate_time(block_prev, timed_pgms)
                if time_s is None:
                    continue  # cannot place without any timing reference

            # Parse picture references
            pictures = _PIC_RE.findall(text)

            # Categorize
            text_lower = text.lower()
            is_breakage = any(kw in text_lower for kw in _BREAKAGE_KW)
            is_wear = any(kw in text_lower for kw in _WEAR_KW) and not is_breakage
            is_ok = any(kw in text_lower for kw in _OK_KW) and not is_breakage and not is_wear
            not_measured = "not measured" in text_lower

            category = (
                "breakage" if is_breakage
                else "wear" if is_wear
                else "ok" if is_ok
                else "not_measured" if not_measured
                else "picture" if pictures
                else "note"
            )

            annotations.append({
                "time_s": round(time_s, 1),
                "label": str(mrow.get("process") or mrow.get("tool_name") or "(unnamed)"),
                "text": text,
                "pictures": [f"P{p}" for p in pictures],
                "category": category,
                "is_breakage": is_breakage,
                "side": side_key,
                "block_prev": _jsonable(mrow.get("block_prev")),
                "pgm_line": _jsonable(pgm),
                "tool_id": _jsonable(mrow.get("tool_id")),
                "tool_name": str(mrow.get("tool_name") or ""),
            })

    annotations.sort(key=lambda a: a["time_s"])
    return annotations


def _interpolate_time(
    pgm: Any,
    timed_pgms: List[Tuple[Any, float]],
) -> Optional[float]:
    """Estimate time for a pgm_line from nearby known pgm→time pairs."""
    if not timed_pgms or pgm is None:
        return None
    try:
        pgm_num = float(pgm)
    except (TypeError, ValueError):
        return None
    before: Optional[Tuple[float, float]] = None
    after: Optional[Tuple[float, float]] = None
    for p, t in timed_pgms:
        if p <= pgm_num:
            before = (p, t)
        if p >= pgm_num and after is None:
            after = (p, t)
    if before and after and before[0] != after[0]:
        frac = (pgm_num - before[0]) / (after[0] - before[0])
        return before[1] + frac * (after[1] - before[1])
    if before:
        return before[1]
    if after:
        return after[1]
    return None


# ── Endpoints ────────────────────────────────────────────────────────────────


@router.get("/channels")
async def list_channels(
    session: str = Query("", description="Session id (empty = union of all sessions)"),
):
    """Return all numeric columns from the merged CSV, flagging curated defaults."""
    try:
        df = _load_merged_df(session)
    except FileNotFoundError as e:
        return {"error": str(e), "channels": []}

    # Only numeric columns
    numeric_cols = sorted(df.select_dtypes(include=["number"]).columns.tolist())

    channels = []
    for col in numeric_cols:
        channels.append({
            "name": col,
            "default": col in _DEFAULT_CHANNELS,
        })

    return {
        "channels": channels,
        "total": len(channels),
        "defaults": _DEFAULT_CHANNELS,
    }


@router.get("/ofs")
async def list_ofs():
    """Return available OF IDs grouped by session."""
    try:
        all_sessions = _load_all_sessions()
    except FileNotFoundError as e:
        return {"error": str(e), "ofs": [], "sessions": []}

    of_col = "UF5-Numero_de_pieza_OF"
    mo_col = "manufacturing_order"

    ofs: List[Dict[str, Any]] = []
    seen_of_session: Set[Tuple[str, str]] = set()

    for sess_id, df in all_sessions.items():
        # Prefer manufacturing_order; fall back to UF5
        col = mo_col if mo_col in df.columns else of_col if of_col in df.columns else None
        if col is None:
            continue
        vals = df[col].dropna()
        vals = vals[vals > 0] if pd.api.types.is_numeric_dtype(vals) else vals
        for of_val in sorted(vals.unique()):
            try:
                of_str = str(int(of_val))
            except (TypeError, ValueError):
                of_str = str(of_val)
            key = (of_str, sess_id)
            if key in seen_of_session:
                continue
            seen_of_session.add(key)
            count = int((vals == of_val).sum())
            ofs.append({
                "id": of_str,
                "session": sess_id,
                "session_label": _get_sessions_registry().get(sess_id, {}).get("label", sess_id),
                "sample_count": count,
            })

    # Session metadata
    sessions = []
    for sess_id, info in _get_sessions_registry().items():
        if sess_id in all_sessions:
            sessions.append({
                "id": sess_id,
                "label": info["label"],
                "date_range": info["date_range"],
                "channels": info["channels"],
                "rows": len(all_sessions[sess_id]),
            })

    return {"ofs": ofs, "sessions": sessions}


@router.get("/waveform")
async def dataset_waveform(
    channels: str = Query("", description="Comma-separated channel names (empty = defaults)"),
    of_id: str = Query("", description="OF number to filter by (empty = all data)"),
    session: str = Query("", description="Session id to load from (empty = auto-detect from OF)"),
    max_points: int = Query(3000, ge=200, le=15000, description="Max points per channel after LTTB"),
    time_min: Optional[float] = Query(None, description="Start of visible window in seconds from OF start (for zoom)"),
    time_max: Optional[float] = Query(None, description="End of visible window in seconds from OF start (for zoom)"),
):
    """Serve downsampled waveform data with process annotation regions.

    When ``time_min`` / ``time_max`` are provided the LTTB downsampling is
    applied only to data within that window, giving full resolution when
    the user zooms in.  Regions are always computed for the full OF.
    """
    try:
        df = _load_merged_df(session)
    except FileNotFoundError as e:
        return {"error": str(e)}

    want = set(c.strip() for c in channels.split(",") if c.strip()) if channels else set(_DEFAULT_CHANNELS)

    # Filter by OF — use manufacturing_order if available, fall back to UF5
    of_col = "manufacturing_order" if "manufacturing_order" in df.columns else "UF5-Numero_de_pieza_OF"
    of_filter_id = ""
    resolved_session = session or ""
    if of_id and of_col in df.columns:
        try:
            of_num = int(of_id)
            df_of = df[df[of_col] == of_num].copy()
            of_filter_id = of_id
            # Resolve session from data if not provided
            if not resolved_session and "_session" in df_of.columns and not df_of.empty:
                resolved_session = str(df_of["_session"].iloc[0])
        except (TypeError, ValueError):
            df_of = df.copy()
    else:
        df_of = df.copy()

    if df_of.empty:
        return {"error": f"No data for OF {of_id}", "channels": []}

    # Compute seconds from start
    if "Date" in df_of.columns and pd.api.types.is_datetime64_any_dtype(df_of["Date"]):
        t0 = df_of["Date"].iloc[0]
        df_of["_seconds"] = (df_of["Date"] - t0).dt.total_seconds()
    else:
        df_of["_seconds"] = np.arange(len(df_of), dtype=float)

    duration_s = float(df_of["_seconds"].iloc[-1] - df_of["_seconds"].iloc[0])
    if not math.isfinite(duration_s):
        duration_s = 0.0

    # Build channels — optionally clip to visible time window before LTTB
    result_channels = []
    for col in sorted(want):
        if col not in df_of.columns:
            continue
        raw_x = df_of["_seconds"].values.astype(float)
        raw_y = df_of[col].values.astype(float)

        # Filter non-finite
        mask = np.isfinite(raw_y)
        raw_x = raw_x[mask]
        raw_y = raw_y[mask]

        if len(raw_x) == 0:
            continue

        # Sort by time
        order = np.argsort(raw_x)
        raw_x = raw_x[order]
        raw_y = raw_y[order]

        # Clip to visible window when zoomed
        if time_min is not None or time_max is not None:
            lo = time_min if time_min is not None else raw_x[0]
            hi = time_max if time_max is not None else raw_x[-1]
            win_mask = (raw_x >= lo) & (raw_x <= hi)
            raw_x = raw_x[win_mask]
            raw_y = raw_y[win_mask]
            if len(raw_x) == 0:
                continue

        # LTTB
        ds_x, ds_y = _lttb_downsample(raw_x, raw_y, max_points)

        result_channels.append({
            "name": col,
            "timestamps": [round(float(v), 1) for v in ds_x],
            "values": [_safe_float(v) for v in ds_y],
        })

    # Determine side from program name
    side = "A"
    if "Cnc_Program_Name_RT" in df_of.columns:
        prg = df_of["Cnc_Program_Name_RT"].dropna()
        if len(prg) > 0:
            first_prog = str(prg.iloc[0]).upper()
            if first_prog.startswith("B"):
                side = "B"

    # Build process regions from milestones
    regions: List[Dict[str, Any]] = []
    annotations: List[Dict[str, Any]] = []
    not_measured = False
    insert_condition = ""
    try:
        milestones = _load_milestones()
        regions = _build_process_regions(df_of, milestones, of_filter_id, side)
        annotations = _extract_annotations(regions, milestones, of_filter_id)

        # Check if this OF was measured — look at the first annotation row
        # (row 2 in xlsx) for "Not measured" vs insert condition
        a_sheet = milestones.get("sheets", {}).get("A")
        if a_sheet and of_filter_id:
            for mrow in a_sheet["rows"][:5]:  # first few rows
                ann_text = mrow.get("of_annotations", {}).get(of_filter_id)
                if ann_text and isinstance(ann_text, str):
                    ann_lower = ann_text.strip().lower()
                    if "not measured" in ann_lower:
                        not_measured = True
                        insert_condition = "not measured"
                        break
                    elif any(kw in ann_lower for kw in ("new inserts", "used inserts")):
                        insert_condition = ann_text.strip()
                        break
    except Exception as e:
        logger.warning("Failed to build process regions: %s", e)

    # Metadata
    sess_info = _get_sessions_registry().get(resolved_session, {})
    metadata: Dict[str, Any] = {
        "of_id": of_filter_id,
        "session": resolved_session,
        "session_label": sess_info.get("label", resolved_session),
        "side": side,
        "total_rows": len(df_of),
        "duration_s": round(duration_s, 1),
        "duration_h": round(duration_s / 3600, 2),
        "not_measured": not_measured,
        "insert_condition": insert_condition,
    }
    if "Cnc_Program_Name_RT" in df_of.columns:
        prgs = df_of["Cnc_Program_Name_RT"].dropna().unique()
        metadata["programs"] = [str(p) for p in prgs]

    return {
        "channels": result_channels,
        "regions": regions,
        "annotations": annotations,
        "metadata": metadata,
    }


@router.get("/milestones")
async def get_milestones(
    side: str = Query("A", description="Sheet side: A or B"),
):
    """Return parsed OFs milestone data for a specific side."""
    try:
        data = _load_milestones()
    except FileNotFoundError as e:
        return {"error": str(e)}

    sheet = data["sheets"].get(side)
    if not sheet:
        available = list(data["sheets"].keys())
        return {"error": f"Side '{side}' not found. Available: {available}"}

    return _jsonable({
        "side": side,
        "name": sheet["name"],
        "total_milestones": sheet["total_milestones"],
        "of_ids": sheet["of_ids"],
        "of_columns": sheet["of_columns"],
        "breakage_summary": sheet["breakage_summary"],
        "tool_summary": sheet["tool_summary"],
        "milestones": sheet["rows"],
    })
